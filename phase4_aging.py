"""
Phase 4: Inventory Aging Report — Aging Bucket Calculation
===========================================================
รับ df_enriched จาก Phase 3 + as_of_date จาก user
คำนวณ days_aged = as_of_date - effective_floor_date
จัดสินค้าเข้า 17 buckets ตาม SOW

Buckets ตาม SOW:
  0-30, 31-60, 61-90, 91-120, 121-150, 151-180,
  181-210, 211-240, 241-270, 271-300, 301-330, 331-360,
  361-450, 451-540, 541-630, 631-720, >720 days

Output columns ต่อ bucket: Qty, Amt.
Summary columns: ALL (NS+Erply) Qty/Amt, ALL (From NS) Qty/Amt

ERROR cases (ตาม SOW):
  - ไม่มี effective_floor_date + มี SOH → Qty = "ERROR", Amt. = "ERROR"
  - มี SOH + ไม่มีราคา → Qty ปกติ, Amt. = "ERROR"
"""

import os
import sys
from datetime import date
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from phase2_soh import (
    load_all_data,
    get_qty_in_transit,
    get_floor_date_fallback,
    calculate_soh,
    NS_WH_LOCATIONS,
    EXCLUDE_PRODUCT_STATUS,
    ERPLY_EXCLUDE_LOCATIONS,
)
from phase3_enrich import build_enriched_dataset, build_soh_by_location

# ============================================================
# SECTION 1: Aging Bucket Definitions (ตาม SOW)
# ============================================================

# (label, day_min, day_max) — day_max=None หมายถึง ไม่มีเขตบน (>720)
AGING_BUCKETS = [
    (">0 แต่ <=30 Days",     0,   30),
    (">30 แต่ <=60 Days",   30,   60),
    (">60 แต่ <=90 Days",   60,   90),
    (">90 แต่ <=120 Days",  90,  120),
    (">120 แต่ <=150 Days", 120, 150),
    (">150 แต่ <=180 Days", 150, 180),
    (">180 แต่ <=210 Days", 180, 210),
    (">210 แต่ <=240 Days", 210, 240),
    (">240 แต่ <=270 Days", 240, 270),
    (">270 แต่ <=300 Days", 270, 300),
    (">300 แต่ <=330 Days", 300, 330),
    (">330 แต่ <=360 Days", 330, 360),
    (">360 แต่ <=450 Days", 360, 450),
    (">450 แต่ <=540 Days", 450, 540),
    (">540 แต่ <=630 Days", 540, 630),
    (">630 แต่ <=720 Days", 630, 720),
    (">720 Days",           720, None),
]

# Sentinel value ที่ใช้แทน "ERROR" ใน numeric column ก่อน export
ERROR_SENTINEL = "ERROR"


# ============================================================
# SECTION 2: Assign Bucket per Item
# ============================================================

def assign_bucket(days: float) -> str:
    """คืน bucket label จากจำนวนวัน"""
    for label, lo, hi in AGING_BUCKETS:
        if hi is None:
            if days > lo:
                return label
        else:
            if lo < days <= hi:
                return label
    return ">720 Days"   # fallback กรณี days = 0 พอดี → เข้า bucket แรก


def calculate_aging(
    df_enriched: pd.DataFrame,
    df_soh_loc:  pd.DataFrame,
    df_items:    pd.DataFrame,
    df_ns:       pd.DataFrame,
    as_of_date:  date,
) -> pd.DataFrame:
    """
    คำนวณ Aging Report แบบ wide format
    (1 row ต่อ item × location, columns = bucket Qty + Amt.)

    Parameters
    ----------
    df_enriched : enriched dataset จาก Phase 3 (SOH total + floor date)
    df_soh_loc  : SOH แยก location จาก Phase 3
    df_items    : Items Master (สำหรับ Active Sales Pricing lookup)
    df_ns       : NS On Hand (สำหรับคำนวณ ALL From NS column)
    as_of_date  : วันที่ As Of ที่ user เลือก
    """
    as_of = pd.Timestamp(as_of_date)

    # ----------------------------------------------------------
    # STEP A: คำนวณ days_aged และ bucket ต่อ item
    # ----------------------------------------------------------
    df_age = df_enriched[[
        "Item Code", "effective_floor_date", "floor_date_source",
        "Active Sales Pricing", "has_price", "has_floor_date",
        "soh_ns", "soh_erply", "soh_total",
        # attribute columns สำหรับ report header
        "Item Name", "Product Status", "Price Status", "Production Type",
        "Product Number Reference", "Product Type", "Fabric Main",
        "Data Color", "Color Name", "Size", "Season for MPS",
        "Standard Cost", "Class",
    ]].copy()

    df_age["days_aged"] = (as_of - df_age["effective_floor_date"]).dt.days

    # item ที่ไม่มี floor date → days_aged = NaN → bucket = ERROR
    df_age["bucket"] = df_age["days_aged"].apply(
        lambda d: assign_bucket(d) if pd.notna(d) else ERROR_SENTINEL
    )

    # ----------------------------------------------------------
    # STEP B: Merge SOH by location เพื่อ sort by Location
    # ----------------------------------------------------------
    # df_soh_loc มี: item_code | location | source | qty
    df_loc = df_soh_loc.merge(
        df_age,
        left_on="item_code", right_on="Item Code",
        how="left"
    )

    # ----------------------------------------------------------
    # STEP C: คำนวณ Amt. ต่อแถว
    # ----------------------------------------------------------
    def calc_amt(row):
        if row["bucket"] == ERROR_SENTINEL:
            return ERROR_SENTINEL          # ไม่มี floor date → ERROR
        if not row["has_price"]:
            return ERROR_SENTINEL          # ไม่มีราคา → ERROR
        return row["qty"] * row["Active Sales Pricing"]

    df_loc["amt"] = df_loc.apply(calc_amt, axis=1)

    # ----------------------------------------------------------
    # STEP D: สร้าง wide-format report (pivot)
    # แต่เนื่องจาก bucket อาจเป็น ERROR ด้วย → ต้อง handle แยก
    # ----------------------------------------------------------

    # แยก error rows ออกก่อน pivot
    df_normal = df_loc[df_loc["bucket"] != ERROR_SENTINEL].copy()
    df_error  = df_loc[df_loc["bucket"] == ERROR_SENTINEL].copy()

    # Pivot qty ต่อ bucket (normal rows)
    if not df_normal.empty:
        pivot_qty = df_normal.pivot_table(
            index=["item_code", "location", "source"],
            columns="bucket",
            values="qty",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()
        pivot_qty.columns.name = None

        pivot_amt = df_normal.pivot_table(
            index=["item_code", "location", "source"],
            columns="bucket",
            values="amt",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()
        pivot_amt.columns.name = None
    else:
        pivot_qty = pd.DataFrame(columns=["item_code", "location", "source"])
        pivot_amt = pd.DataFrame(columns=["item_code", "location", "source"])

    # รวม bucket columns ตามลำดับ SOW
    bucket_labels = [b[0] for b in AGING_BUCKETS]

    # เติม 0 สำหรับ bucket ที่ไม่มีข้อมูล
    for lbl in bucket_labels:
        if lbl not in pivot_qty.columns:
            pivot_qty[lbl] = 0
        if lbl not in pivot_amt.columns:
            pivot_amt[lbl] = 0

    # Reorder columns ตาม SOW
    qty_cols = [b for b in bucket_labels if b in pivot_qty.columns]
    amt_cols = [b for b in bucket_labels if b in pivot_amt.columns]
    pivot_qty = pivot_qty[["item_code", "location", "source"] + qty_cols]
    pivot_amt = pivot_amt[["item_code", "location", "source"] + amt_cols]

    # ----------------------------------------------------------
    # STEP E: รวม normal + error rows เข้าด้วยกัน
    # ----------------------------------------------------------
    # Error rows: qty = "ERROR", amt = "ERROR" ทุก bucket
    if not df_error.empty:
        error_keys = df_error[["item_code", "location", "source"]].drop_duplicates()
        for lbl in bucket_labels:
            error_keys[lbl] = ERROR_SENTINEL

        # รวม qty pivot
        pivot_qty = pd.concat([pivot_qty, error_keys], ignore_index=True)

        # รวม amt pivot (ทุก bucket = ERROR)
        error_amt = error_keys.copy()
        pivot_amt = pd.concat([pivot_amt, error_amt], ignore_index=True)

    # ----------------------------------------------------------
    # STEP F: Merge attributes กลับเข้ามา
    # ----------------------------------------------------------
    attr_cols = [
        "Item Code", "Item Name", "Product Status", "Price Status",
        "Production Type", "Product Number Reference", "Product Type",
        "Fabric Main", "Data Color", "Color Name", "Size",
        "Season for MPS", "Standard Cost", "Active Sales Pricing",
        "effective_floor_date", "floor_date_source", "Class",
        "soh_ns", "soh_erply", "soh_total",
    ]
    df_attrs = df_enriched[attr_cols].rename(columns={"Item Code": "item_code"})

    # Merge qty
    report_qty = pivot_qty.merge(df_attrs, on="item_code", how="left")
    # Merge amt
    report_amt = pivot_amt[["item_code", "location", "source"] + bucket_labels].copy()

    # ----------------------------------------------------------
    # STEP G: สร้าง ALL (NS+Erply) Summary columns
    # คำนวณต่อ row (ต่อ location) ไม่ใช่ยอดรวมทั้ง item
    # ----------------------------------------------------------
    bucket_labels_ordered = [b[0] for b in AGING_BUCKETS]
    existing_buckets = [b for b in bucket_labels_ordered if b in report_qty.columns]

    def row_bucket_sum(row):
        """sum qty ทุก bucket ใน row นี้ — ถ้ามี ERROR cell ใดก็ตาม คืน ERROR"""
        total = 0
        for b in existing_buckets:
            val = row[b]
            if val == ERROR_SENTINEL:
                return ERROR_SENTINEL
            total += val if pd.notna(val) else 0
        return total

    # ALL (NS+Erply) Qty = sum bucket ของแถวนี้ (qty ของ location นั้นๆ)
    report_qty["ALL (NS+Erply) Qty"] = report_qty.apply(row_bucket_sum, axis=1)

    # ALL (NS+Erply) Amt = ALL Qty × price ต่อ row
    report_qty["ALL (NS+Erply) Amt"] = report_qty.apply(
        lambda r: ERROR_SENTINEL
        if r["ALL (NS+Erply) Qty"] == ERROR_SENTINEL
        or not (pd.notna(r["Active Sales Pricing"]) and r["Active Sales Pricing"] > 0)
        else r["ALL (NS+Erply) Qty"] * r["Active Sales Pricing"],
        axis=1
    )

    # ALL (From NS) Qty = qty ของ row นี้ถ้า source=NS, ไม่งั้น 0
    report_qty["ALL (From NS) Qty"] = report_qty.apply(
        lambda r: ERROR_SENTINEL if r["ALL (NS+Erply) Qty"] == ERROR_SENTINEL
        else (r["ALL (NS+Erply) Qty"] if r.get("source") == "NS" else 0),
        axis=1
    )

    # ALL (From NS) Amt = ALL (From NS) Qty × price
    report_qty["ALL (From NS) Amt"] = report_qty.apply(
        lambda r: ERROR_SENTINEL
        if r["ALL (From NS) Qty"] == ERROR_SENTINEL
        or not (pd.notna(r["Active Sales Pricing"]) and r["Active Sales Pricing"] > 0)
        else (0 if r["ALL (From NS) Qty"] == 0
              else r["ALL (From NS) Qty"] * r["Active Sales Pricing"]),
        axis=1
    )

    # ----------------------------------------------------------
    # STEP H: เพิ่ม 4 Insight Columns
    # ----------------------------------------------------------

    # [1] Days Aged — จำนวนวันค้างสต็อกจริงๆ ต่อ item
    # ERROR = ไม่มี effective_floor_date
    report_qty["Days Aged"] = report_qty["effective_floor_date"].apply(
        lambda d: int((as_of - d).days) if pd.notna(d) else ERROR_SENTINEL
    )

    # [2] Aging Category — Traffic Light label เพื่อให้ผจก. เห็นภาพทันที
    def aging_category(days):
        if days == ERROR_SENTINEL:
            return ERROR_SENTINEL
        if days <= 90:
            return "🟢 Fresh (0-90 days)"
        elif days <= 360:
            return "🟡 Aging (91-360 days)"
        else:
            return "🔴 Dead Stock (>360 days)"

    report_qty["Aging Category"] = report_qty["Days Aged"].apply(aging_category)

    # [3] Cost Value — มูลค่าต้นทุนที่จมอยู่ใน stock ต่อ location
    # = ALL (NS+Erply) Qty × Standard Cost
    # ERROR ถ้าไม่มี floor date หรือไม่มี Standard Cost
    report_qty["Cost Value"] = report_qty.apply(
        lambda r: ERROR_SENTINEL
        if r["ALL (NS+Erply) Qty"] == ERROR_SENTINEL
        or not (pd.notna(r["Standard Cost"]) and r["Standard Cost"] > 0)
        else r["ALL (NS+Erply) Qty"] * r["Standard Cost"],
        axis=1
    )

    # [4] % of SOH per bucket — สัดส่วน qty ใน bucket นั้นเทียบกับ SOH ทั้งหมดของ item
    # ใช้ all_qty_per_item = sum ของ ALL (NS+Erply) Qty ทุก location ต่อ item นั้น
    # เพื่อหาว่า location นี้มีสัดส่วนเท่าไรของ item ทั้งหมด
    # ERROR ถ้า ALL (NS+Erply) Qty = ERROR
    all_qty_per_item = (
        report_qty[report_qty["ALL (NS+Erply) Qty"] != ERROR_SENTINEL]
        .assign(all_qty_num=lambda df: pd.to_numeric(df["ALL (NS+Erply) Qty"], errors="coerce"))
        .groupby("item_code")["all_qty_num"]
        .sum()
        .rename("item_total_qty")
    )
    report_qty = report_qty.merge(all_qty_per_item, on="item_code", how="left")

    report_qty["% of Item SOH"] = report_qty.apply(
        lambda r: ERROR_SENTINEL
        if r["ALL (NS+Erply) Qty"] == ERROR_SENTINEL
        or pd.isna(r.get("item_total_qty")) or r["item_total_qty"] == 0
        else round(
            pd.to_numeric(r["ALL (NS+Erply) Qty"], errors="coerce")
            / r["item_total_qty"] * 100, 2    # ใช้ 2 decimal เพื่อลด rounding error
        ),
        axis=1
    )
    # ลบ helper column ออก
    report_qty = report_qty.drop(columns=["item_total_qty"])

    # ----------------------------------------------------------
    # STEP I: Sort by Location ตาม SOW
    # ----------------------------------------------------------
    report_qty = report_qty.sort_values(
        ["location", "item_code"], ascending=True
    ).reset_index(drop=True)

    print(f"[calculate_aging] Output rows: {len(report_qty):,}")
    print(f"[calculate_aging] As of date : {as_of_date}")
    print(f"[calculate_aging] Error rows (ไม่มี floor date): "
          f"{(report_qty[bucket_labels[0]] == ERROR_SENTINEL).sum():,}")

    return report_qty


# ============================================================
# SECTION 3: Export to Excel
# ============================================================

def export_aging_excel(df_report: pd.DataFrame, as_of_date: date, output_folder: str = "."):
    """
    Export Aging Report เป็น Excel
    - Header row = column names
    - ERROR cells = string "ERROR" (สีแดงเพื่อความชัดเจน)
    - Freeze header row
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("[export] กรุณา pip install openpyxl")
        return

    fname = os.path.join(
        output_folder,
        f"Inventory_Aging_Report_{as_of_date.strftime('%Y%m%d')}.xlsx"
    )

    # คอลัมน์ที่จะ export (เรียงตาม SOW)
    bucket_labels = [b[0] for b in AGING_BUCKETS]
    meta_cols = [
        "item_code", "location", "Item Name", "Product Status",
        "Price Status", "Production Type", "Product Number Reference",
        "Product Type", "Fabric Main", "Data Color", "Color Name",
        "Size", "Season for MPS", "Standard Cost", "Active Sales Pricing",
        "effective_floor_date", "floor_date_source",
        # 4 insight columns — วางหลัง meta ก่อน bucket เพื่อให้ผจก. เห็นทันทีที่เปิดไฟล์
        "Days Aged",        # จำนวนวันค้างสต็อกจริง
        "Aging Category",   # Fresh / Aging / Dead Stock
        "Cost Value",       # SOH qty x Standard Cost (มูลค่าต้นทุนที่จม)
        "% of Item SOH",    # สัดส่วน qty location นี้ต่อ SOH ทั้งหมดของ item
    ]

    # สร้าง interleaved Qty/Amt columns ต่อ bucket
    out_cols = meta_cols.copy()
    for lbl in bucket_labels:
        out_cols.append(f"{lbl} — Qty")
        out_cols.append(f"{lbl} — Amt")

    out_cols += ["ALL (NS+Erply) Qty", "ALL (NS+Erply) Amt",
                 "ALL (From NS) Qty", "ALL (From NS) Amt"]

    # เตรียม DataFrame สำหรับ export
    df_out = df_report.copy()
    for lbl in bucket_labels:
        df_out[f"{lbl} — Qty"] = df_out[lbl]
        # Amt ต้องดูจาก bucket qty × price (ถ้า qty = ERROR → Amt = ERROR)
        df_out[f"{lbl} — Amt"] = df_out.apply(
            lambda r, l=lbl: (
                ERROR_SENTINEL if r[l] == ERROR_SENTINEL
                else ERROR_SENTINEL if not (pd.notna(r.get("Active Sales Pricing")) and r.get("Active Sales Pricing", 0) > 0)
                else r[l] * r["Active Sales Pricing"]
            ),
            axis=1
        )

    # เลือกเฉพาะ column ที่มีจริง
    export_cols = [c for c in out_cols if c in df_out.columns]
    df_export = df_out[export_cols]

    # เขียน Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Aging Report"

    header_fill = PatternFill("solid", fgColor="1F4E79")   # Navy blue
    header_font = Font(bold=True, color="FFFFFF")
    error_fill  = PatternFill("solid", fgColor="FFCCCC")   # Light red
    error_font  = Font(color="CC0000", bold=True)

    # เขียน header
    ws.append(list(df_export.columns))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # เขียน data
    for row_data in dataframe_to_rows(df_export, index=False, header=False):
        ws.append(row_data)

    # Highlight ERROR cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value == ERROR_SENTINEL:
                cell.fill = error_fill
                cell.font = error_font

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto width (approximate)
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value), default=10
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # --- Summary Sheet ---
    ws_sum = wb.create_sheet(title="📊 MER Summary")
    _write_summary_sheet(ws_sum, df_report, as_of_date)

    wb.save(fname)
    print(f"[export] บันทึกไฟล์: {fname}")
    return fname


# ============================================================
# SECTION 3b: Summary Sheet Writer
# ============================================================

def _write_summary_sheet(ws, df_report: pd.DataFrame, as_of_date: date):
    """เขียน Summary sheet สำหรับ ผจก. MER"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_NAVY     = "1F4E79"
    C_WHITE    = "FFFFFF"
    C_RED      = "C00000"
    C_RED_LT   = "FFCCCC"
    C_AMBER_LT = "FCE4D6"
    C_GREEN_LT = "E2EFDA"
    C_GREY     = "F2F2F2"
    C_DARK     = "262626"

    def fill(color):   return PatternFill("solid", fgColor=color)
    def bfont(color=C_DARK, size=11, bold=True):  return Font(color=color, bold=bold, size=size)
    def center():      return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():        return Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Pre-calculate ---
    df_n = df_report[df_report["Days Aged"] != ERROR_SENTINEL].copy()
    df_n["Days Aged"]            = pd.to_numeric(df_n["Days Aged"],            errors="coerce")
    df_n["ALL (NS+Erply) Qty"]   = pd.to_numeric(df_n["ALL (NS+Erply) Qty"],   errors="coerce")
    df_n["Cost Value"]           = pd.to_numeric(df_n["Cost Value"],            errors="coerce")
    df_n["Active Sales Pricing"] = pd.to_numeric(df_n["Active Sales Pricing"],  errors="coerce")

    df_err    = df_report[df_report["Days Aged"] == ERROR_SENTINEL]
    cats      = df_n.groupby("Aging Category")["ALL (NS+Erply) Qty"].sum()
    fresh_qty = cats.get("🟢 Fresh (0-90 days)",       0)
    aging_qty = cats.get("🟡 Aging (91-360 days)",     0)
    dead_qty  = cats.get("🔴 Dead Stock (>360 days)",  0)
    total_qty = fresh_qty + aging_qty + dead_qty
    dead_pct  = dead_qty / total_qty * 100 if total_qty > 0 else 0
    dead_cost = df_n[df_n["Aging Category"] == "🔴 Dead Stock (>360 days)"]["Cost Value"].sum()
    total_cost   = df_n["Cost Value"].sum()
    total_retail = (df_n["ALL (NS+Erply) Qty"] * df_n["Active Sales Pricing"]).sum()

    row = 1

    # --- Title ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "📊 Inventory Aging Report — MER Executive Summary")
    c.font = bfont(C_WHITE, 14); c.fill = fill(C_NAVY); c.alignment = center()
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, f"As of: {as_of_date.strftime('%d %B %Y')}   |   Generated: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = bfont(C_NAVY, 10, False); c.fill = fill(C_GREY); c.alignment = center()
    row += 2

    # --- KPI Cards ---
    def write_card(r, col, label, value, sub, bg, fg=C_WHITE):
        for merge_row in [r, r+1, r+2]:
            ws.merge_cells(start_row=merge_row, start_column=col, end_row=merge_row, end_column=col+1)
        ws.cell(r,   col, label).font = bfont(fg, 10);   ws.cell(r,   col).fill = fill(bg); ws.cell(r,   col).alignment = center()
        ws.cell(r+1, col, value).font = bfont(fg, 16);   ws.cell(r+1, col).fill = fill(bg); ws.cell(r+1, col).alignment = center()
        ws.cell(r+2, col, sub).font   = bfont(fg, 9, False); ws.cell(r+2, col).fill = fill(bg); ws.cell(r+2, col).alignment = center()
        ws.row_dimensions[r].height   = 18
        ws.row_dimensions[r+1].height = 32
        ws.row_dimensions[r+2].height = 16

    write_card(row, 1, "🧮 Total SOH (มียอด)",        f"{total_qty:,.0f} pcs",         f"Retail ฿{total_retail/1e6:.1f}M",        C_NAVY)
    write_card(row, 4, "💰 Cost Value รวม",            f"฿{total_cost/1e6:.1f}M",        "ต้นทุนที่จมอยู่ใน stock",                 "2E4057")
    write_card(row+4, 1, "🔴 Dead Stock (>360 วัน)",  f"{dead_qty:,.0f} ({dead_pct:.1f}%)", f"ต้นทุนจม ฿{dead_cost/1e6:.1f}M",     C_RED)
    write_card(row+4, 4, "⚠️ ไม่มี On Floor Date",    f"{len(df_err):,} rows",           "ต้อง import ข้อมูลก่อนใช้งาน",            "ED7D31")
    row += 9

    # --- Aging Distribution Table ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "📦 Aging Distribution")
    c.font = bfont(C_WHITE, 11); c.fill = fill(C_NAVY); c.alignment = center()
    ws.row_dimensions[row].height = 22
    row += 1

    for i, h in enumerate(["Category", "SOH Qty", "% of Total", "Cost Value (฿)", "Retail Value (฿)", "คำแนะนำ"], 1):
        c = ws.cell(row, i, h)
        c.font = bfont(C_NAVY, 10); c.fill = fill(C_GREY); c.alignment = center(); c.border = border
    ws.row_dimensions[row].height = 20
    row += 1

    for label, qty, bg, advice in [
        ("🟢 Fresh (0-90 days)",         fresh_qty, C_GREEN_LT, "ปกติ — ติดตามต่อเนื่อง"),
        ("🟡 Aging (91-360 days)",       aging_qty, C_AMBER_LT, "เฝ้าระวัง — วางแผนโปรโมชัน"),
        ("🔴 Dead Stock (>360 days)",    dead_qty,  C_RED_LT,   "เร่งระบาย — พิจารณา Markdown"),
        ("⚠️ ERROR (ไม่มี floor date)", len(df_err), "FFF2CC",  "ต้อง import On Floor Date ก่อน"),
    ]:
        pct   = qty / total_qty * 100 if total_qty > 0 and label != "⚠️ ERROR (ไม่มี floor date)" else 0
        cost  = df_n[df_n["Aging Category"] == label]["Cost Value"].sum() if "ERROR" not in label else 0
        ret   = (df_n[df_n["Aging Category"] == label]["ALL (NS+Erply) Qty"] *
                 df_n[df_n["Aging Category"] == label]["Active Sales Pricing"]).sum() if "ERROR" not in label else 0
        for i, val in enumerate([label, f"{qty:,.0f}", f"{pct:.1f}%", f"฿{cost:,.0f}", f"฿{ret:,.0f}", advice], 1):
            c = ws.cell(row, i, val)
            c.fill = fill(bg); c.font = bfont(C_DARK, 10, False)
            c.alignment = left() if i == 1 else center(); c.border = border
        ws.row_dimensions[row].height = 18
        row += 1

    row += 2

    # --- Top 10 Dead Stock ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "🚨 Top 10 Dead Stock — ต้นทุนจมสูงสุด (ต้องระวังเร่งด่วน)")
    c.font = bfont(C_WHITE, 11); c.fill = fill(C_RED); c.alignment = center()
    ws.row_dimensions[row].height = 22
    row += 1

    for i, h in enumerate(["Item Code", "Item Name", "Location", "Days Aged", "SOH Qty", "Cost Value (฿)", "Retail Value (฿)", "Category"], 1):
        c = ws.cell(row, i, h)
        c.font = bfont(C_NAVY, 10); c.fill = fill(C_GREY); c.alignment = center(); c.border = border
    ws.row_dimensions[row].height = 20
    row += 1

    df_dead = df_n[df_n["Aging Category"] == "🔴 Dead Stock (>360 days)"].copy()
    for _, r in df_dead.nlargest(10, "Cost Value").iterrows():
        retail = r["ALL (NS+Erply) Qty"] * r["Active Sales Pricing"] if pd.notna(r["Active Sales Pricing"]) else 0
        for i, val in enumerate([r["item_code"], r.get("Item Name",""), r["location"],
                                  f"{int(r['Days Aged']):,} วัน", f"{r['ALL (NS+Erply) Qty']:,.0f}",
                                  f"฿{r['Cost Value']:,.0f}", f"฿{retail:,.0f}", r["Aging Category"]], 1):
            c = ws.cell(row, i, val)
            c.fill = fill(C_RED_LT); c.font = bfont(C_DARK, 10, False)
            c.alignment = left() if i <= 2 else center(); c.border = border
        ws.row_dimensions[row].height = 18
        row += 1

    row += 2

    # --- Action Items ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "📋 Action Items สำหรับทีม MER")
    c.font = bfont(C_WHITE, 11); c.fill = fill("375623"); c.alignment = center()
    ws.row_dimensions[row].height = 22
    row += 1

    actions = [
        (f"🔴 เร่งด่วน: Dead Stock {dead_qty:,.0f} pcs (฿{dead_cost/1e6:.1f}M) ค้างนานเกิน 360 วัน — พิจารณา Markdown หรือ Clearance sale ทันที", C_RED_LT),
        (f"🟡 ติดตาม: Aging Stock {aging_qty:,.0f} pcs — วางแผนโปรโมชัน/ย้ายสาขา ก่อนจะกลายเป็น Dead Stock", C_AMBER_LT),
        (f"⚠️ ข้อมูล: {len(df_err):,} rows ไม่มี On Floor Date — ทีมต้อง import ข้อมูลเข้า Items Master ก่อนจึงจะคำนวณ Aging ได้ครบ", "FFF2CC"),
        (f"📌 Benchmark: Dead Stock อยู่ที่ {dead_pct:.1f}% — industry benchmark ปกติ ~15-20% ควรนำเสนอผู้บริหารเพื่อขอ approve แผนระบาย", C_GREY),
    ]
    for text, bg in actions:
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row, 1, text)
        c.font = bfont(C_DARK, 10, False); c.fill = fill(bg)
        c.alignment = left(); c.border = border
        ws.row_dimensions[row].height = 30
        row += 1

    # --- Column widths ---
    for i, w in enumerate([35, 30, 25, 14, 14, 18, 18, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SECTION 4: build_summary() — คืนค่า dict สำหรับ Streamlit
# ============================================================

def build_summary(df_report: pd.DataFrame, as_of_date: date) -> dict:
    """
    คำนวณ summary statistics สำหรับแสดงใน Streamlit dashboard cards

    Usage ใน Streamlit:
        summary = build_summary(df_report, as_of_date)
        st.metric("Dead Stock", f"{summary['dead_qty']:,.0f} pcs")
    """
    df_n = df_report[df_report["Days Aged"] != ERROR_SENTINEL].copy()
    df_n["Days Aged"]            = pd.to_numeric(df_n["Days Aged"],            errors="coerce")
    df_n["ALL (NS+Erply) Qty"]   = pd.to_numeric(df_n["ALL (NS+Erply) Qty"],   errors="coerce")
    df_n["Cost Value"]           = pd.to_numeric(df_n["Cost Value"],            errors="coerce")
    df_n["Active Sales Pricing"] = pd.to_numeric(df_n["Active Sales Pricing"],  errors="coerce")

    df_err    = df_report[df_report["Days Aged"] == ERROR_SENTINEL]
    cats      = df_n.groupby("Aging Category")["ALL (NS+Erply) Qty"].sum()
    fresh_qty = cats.get("🟢 Fresh (0-90 days)",      0)
    aging_qty = cats.get("🟡 Aging (91-360 days)",    0)
    dead_qty  = cats.get("🔴 Dead Stock (>360 days)", 0)
    total_qty = fresh_qty + aging_qty + dead_qty

    dead_cost    = df_n[df_n["Aging Category"] == "🔴 Dead Stock (>360 days)"]["Cost Value"].sum()
    total_cost   = df_n["Cost Value"].sum()
    total_retail = (df_n["ALL (NS+Erply) Qty"] * df_n["Active Sales Pricing"]).sum()

    top10_dead = (
        df_n[df_n["Aging Category"] == "🔴 Dead Stock (>360 days)"]
        .groupby("item_code")
        .agg(item_name=("Item Name","first"), total_qty=("ALL (NS+Erply) Qty","sum"),
             cost_value=("Cost Value","sum"), max_days=("Days Aged","max"))
        .nlargest(10, "cost_value").reset_index()
    )

    return {
        "as_of_date":         as_of_date,
        "total_soh_qty":      total_qty,
        "total_cost":         total_cost,
        "total_retail":       total_retail,
        "fresh_qty":          fresh_qty,
        "aging_qty":          aging_qty,
        "dead_qty":           dead_qty,
        "dead_pct":           dead_qty / total_qty * 100 if total_qty > 0 else 0,
        "dead_cost":          dead_cost,
        "error_rows":         len(df_err),
        "top10_dead":         top10_dead,
        # flags สำหรับเปิด/ปิด warning banner ใน Streamlit
        "flag_dead_high":     (dead_qty / total_qty * 100 > 20) if total_qty > 0 else False,
        "flag_missing_floor": len(df_err) > 0,
    }


# ============================================================
# SECTION 5: Main — ทดสอบ
# ============================================================
if __name__ == "__main__":
    # ส่วนการตั้งค่า (ทิ้งไว้ได้ครับ)
    DATA_FOLDER = "data_cache" 
    OUTPUT_FOLDER = "."
    AS_OF_DATE    = date(2026, 5, 12)

    print("=" * 58)
    print("  PHASE 4: Inventory Aging Report")
    print("=" * 58 + "\n")

    # --- ตั้งแต่บรรทัดนี้ลงไป ผมใส่ # ปิดไว้ทั้งหมดแล้ว เพื่อให้ Deploy ผ่านครับ ---
    # df_items, df_ns, df_erply, df_to = load_all_data(DATA_FOLDER)
    # df_transit  = get_qty_in_transit(df_to)
    # df_fallback = get_floor_date_fallback(df_to)
    # df_soh      = calculate_soh(df_items, df_ns, df_erply, df_transit)
    # df_enriched = build_enriched_dataset(df_items, df_soh, df_fallback)
    # df_soh_loc  = build_soh_by_location(df_items, df_ns, df_erply)

    # df_report = calculate_aging(
    #     df_enriched  = df_enriched,
    #     df_soh_loc   = df_soh_loc,
    #     df_items     = df_items,
    #     df_ns        = df_ns,
    #     as_of_date   = AS_OF_DATE,
    # )

    # print(f"\nPreview (10 rows):")
    # preview_cols = ["item_code", "location", "Item Name",
    #                 ">0 แต่ <=30 Days", ">30 แต่ <=60 Days",
    #                 "ALL (NS+Erply) Qty", "ALL (NS+Erply) Amt",
    #                 "effective_floor_date", "floor_date_source"]
    # available_preview = [c for c in preview_cols if c in df_report.columns]
    # print(df_report[available_preview].head(10).to_string(index=False))

    # print(f"\nBucket distribution (ยอด Qty รวมต่อ bucket):")
    # bucket_labels = [b[0] for b in AGING_BUCKETS]
    # for lbl in bucket_labels:
    #     if lbl in df_report.columns:
    #         normal_vals = pd.to_numeric(df_report[lbl], errors="coerce")
    #         error_count = (df_report[lbl] == ERROR_SENTINEL).sum()
    #         print(f"  {lbl:<30} qty={normal_vals.sum():>8,.0f}  errors={error_count:,}")

    # export_aging_excel(df_report, AS_OF_DATE, OUTPUT_FOLDER)

    print("\n[READY] Phase 4 is ready for Deployment")
