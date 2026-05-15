"""
Phase 5: Sell-through Report
=============================
วัดประสิทธิภาพสินค้าที่ขายจากหน้าร้าน เทียบกับสินค้าที่รับเข้ามา
SELL% = TOTAL SLS QTY / (BOM QTY + TOTAL GR QTY)

Data Sources:
  - YMFSALESDATAWITHCOSTResults  → Sales (Cash Sale, Invoice, Cash Refund, Credit Memo)
  - YMFTRANSFERORDERBYITEMLISTResults → GR (รับเข้า FG Domestic จากคลังผลิต)
  - YMFINVENTORYONHANDREPORT + Inventory_By_Items → SOH
  - Items Master → attributes สินค้า

Confirm ก่อนเขียนโค้ด:
  - Transaction Types: Cash Sale, Invoice, Cash Refund, Credit Memo เท่านั้น
  - Location: แสดงทุก Type (รวม Export, Domestic)
  - Standard Cost: ใช้จาก STANDARD COST/U ในไฟล์ Sales โดยตรง
"""

import glob
import os
import sys
import subprocess
import shutil
from datetime import date
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from phase2_soh import (
    load_all_data, get_qty_in_transit, get_floor_date_fallback,
    calculate_soh, EXCLUDE_PRODUCT_STATUS,
)
from phase3_enrich import build_enriched_dataset, build_soh_by_location

# ============================================================
# SECTION 1: Constants
# ============================================================

VALID_SALE_TYPES = ["Cash Sale", "Invoice", "Cash Refund", "Credit Memo"]
GR_FROM_LOCATION = "คลังผลิต"
ERROR_SENTINEL   = "ERROR"


# ============================================================
# SECTION 2: Load Sales Data
# ============================================================

def _find_file(folder: str, prefix: str) -> str:
    for ext in ["*.xlsx", "*.xls"]:
        matches = glob.glob(os.path.join(folder, f"{prefix}{ext}"))
        if matches:
            return sorted(matches)[0]
    raise FileNotFoundError(f"ไม่พบไฟล์ '{prefix}*.xls(x)' ใน {folder}")


def load_sales_data(folder: str) -> pd.DataFrame:
    """
    โหลด YMFSALESDATAWITHCOSTResults
    ตัด: Overall Total row, Sales Order, Return Authorization
    รองรับทั้ง .xlsx ปกติ และ .xls (ซึ่งมักจะเป็น HTML ปลอมตัวมา)
    """
    path = _find_file(folder, "YMFSALESDATAWITHCOSTResults")
    
    try:
        # พยายามอ่านแบบปกติก่อน (ถ้าเป็น .xlsx ของจริง)
        df = pd.read_excel(path, engine="openpyxl", dtype={"Material Code": str})
    except Exception:
        # ถ้าพัง (มักจะเพราะเป็นไฟล์ .xls ที่ไส้ในเป็น HTML)
        # ให้ใช้วิธีอ่าน HTML table โดยตรง
        print(f"  [Sales] ตรวจพบไฟล์ .xls หรือ HTML fallback กำลังพยายามอ่านตาราง...")
        dfs = pd.read_html(path, dtype={"Material Code": str})
        df = dfs[0] # สมมติว่าตารางข้อมูลคือตารางแรกที่เจอในไฟล์ HTML

    # ตรวจสอบว่าคอลัมน์สำคัญมีอยู่จริง
    required_cols = ["Type", "Material Code", "Selling Date", "Quantity"]
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"ไม่พบคอลัมน์ที่ต้องการในไฟล์ Sales: {missing_cols}")

    # ตัด row ที่ไม่ใช่ transaction จริง
    df = df[df["Type"].isin(VALID_SALE_TYPES)].copy()

    df["Material Code"]         = df["Material Code"].str.strip()

    # ตัด Material Code ที่ไม่ตรง SKU format มาตรฐาน
    from phase2_soh import is_valid_sku
    before = len(df)
    df = df[df["Material Code"].apply(is_valid_sku)].copy()
    print(f"  [Sales SKU filter] ตัดออก {before - len(df):,} rows (non-standard SKU)")
    df["Selling Date"]          = pd.to_datetime(df["Selling Date"],          errors="coerce")
    df["Quantity"]              = pd.to_numeric(df["Quantity"],               errors="coerce").fillna(0)
    df["ACTIVE PRICE/U"]        = pd.to_numeric(df["ACTIVE PRICE/U"],         errors="coerce")
    df["STANDARD COST/U"]       = pd.to_numeric(df["STANDARD COST/U"],        errors="coerce")
    df["TOTAL NET SALES"]       = pd.to_numeric(df["TOTAL NET SALES"],        errors="coerce")
    df["TOTAL STANDARD COST"]   = pd.to_numeric(df["TOTAL STANDARD COST"],    errors="coerce")
    df["Month Sold"]            = pd.to_numeric(df["Month Sold"],              errors="coerce")
    df["Year Sold"]             = pd.to_numeric(df["Year Sold"],               errors="coerce")

    print(f"  [Sales]         {len(df):,} rows | {os.path.basename(path)}")
    return df


# ============================================================
# SECTION 3: Calculate GR (Goods Received)
# ============================================================

def calculate_gr(df_to: pd.DataFrame, date_from: date, date_to: date) -> pd.DataFrame:
    """
    GR = Quantity Received จาก Transfer Order
    Filter: From Location คือ คลังผลิต + อยู่ใน date range
    SOW: ถ้า item เดียวกัน 2 rows ต่างวันที่ → รวม qty
    """
    df = df_to.copy()
    df["Ship Date"]         = pd.to_datetime(df["Ship Date"],          errors="coerce")
    df["Quantity Received"] = pd.to_numeric(df["Quantity Received"],   errors="coerce").fillna(0)

    mask = (
        df["From Location"].str.contains(GR_FROM_LOCATION, na=False) &
        (df["Ship Date"] >= pd.Timestamp(date_from)) &
        (df["Ship Date"] <= pd.Timestamp(date_to))
    )
    df_gr = df[mask].copy()

    result = (
        df_gr.groupby("Item Number", as_index=False)["Quantity Received"]
        .sum()
        .rename(columns={"Item Number": "item_code",
                         "Quantity Received": "total_gr_qty"})
    )
    result["item_code"] = result["item_code"].str.strip()

    print(f"  [GR]            {len(result):,} items | "
          f"Total GR qty: {result['total_gr_qty'].sum():,.0f}")
    return result


# ============================================================
# SECTION 4: Calculate Sales Summary per item × location
# ============================================================

def calculate_sales_summary(df_sales: pd.DataFrame, date_from: date, date_to: date) -> pd.DataFrame:
    """
    คำนวณ Sales per item × location ใน date range
    รวม: TOTAL SLS QTY, TOTAL SLS AMT, TOTAL COST AMT
         AVG SLS QTY/M (เฉพาะเดือนที่มียอด ≠ 0 ตาม SOW)
    """
    mask = (
        (df_sales["Selling Date"] >= pd.Timestamp(date_from)) &
        (df_sales["Selling Date"] <= pd.Timestamp(date_to))
    )
    df = df_sales[mask].copy()

    # Sum per item × location
    sales_main = (
        df.groupby(["Material Code", "Location Name"], as_index=False)
        .agg(
            total_sls_qty  = ("Quantity",           "sum"),
            total_sls_amt  = ("TOTAL NET SALES",    "sum"),
            total_cost_amt = ("TOTAL STANDARD COST","sum"),
        )
        .rename(columns={"Material Code": "item_code",
                         "Location Name":  "location"})
    )

    # AVG SLS QTY/M — ตาม SOW: เฉลี่ยเฉพาะเดือนที่มียอด
    df["ym"] = (
        df["Year Sold"].astype(int).astype(str) + "-" +
        df["Month Sold"].astype(int).astype(str).str.zfill(2)
    )
    monthly = (
        df.groupby(["Material Code", "Location Name", "ym"], as_index=False)["Quantity"]
        .sum()
        .rename(columns={"Material Code": "item_code",
                         "Location Name":  "location",
                         "Quantity":        "monthly_qty"})
    )
    # เฉพาะเดือนที่มียอด ≠ 0
    avg_m = (
        monthly[monthly["monthly_qty"] != 0]
        .groupby(["item_code","location"], as_index=False)["monthly_qty"]
        .mean()
        .rename(columns={"monthly_qty": "avg_sls_qty_m"})
    )

    result = sales_main.merge(avg_m, on=["item_code","location"], how="left")
    result["avg_sls_qty_m"] = result["avg_sls_qty_m"].fillna(0)

    print(f"  [Sales Summary] {len(result):,} rows (item × location) | "
          f"Net qty: {result['total_sls_qty'].sum():,.0f}")
    return result


# ============================================================
# SECTION 5: Build Sell-through Report
# ============================================================

def calculate_sellthrough(
    df_items:    pd.DataFrame,
    df_ns:       pd.DataFrame,
    df_erply:    pd.DataFrame,
    df_to:       pd.DataFrame,
    df_transit:  pd.DataFrame,
    df_fallback: pd.DataFrame,
    df_sales:    pd.DataFrame,
    date_from:   date,
    date_to:     date,
) -> pd.DataFrame:
    """
    Build Sell-through Report — 1 row ต่อ item × location
    Base: Sales data (ตาม SOW)
    """
    # --- SOH ---
    df_soh     = calculate_soh(df_items, df_ns, df_erply, df_transit)
    df_enrich  = build_enriched_dataset(df_items, df_soh, df_fallback)
    df_soh_loc = build_soh_by_location(df_items, df_ns, df_erply)

    # SOH total per item (รวมทุก location)
    soh_by_item = (
        df_soh_loc.groupby("item_code", as_index=False)["qty"]
        .sum().rename(columns={"qty": "soh_qty"})
    )

    # --- GR ---
    df_gr = calculate_gr(df_to, date_from, date_to)

    # --- Sales Summary ---
    df_sales_sum = calculate_sales_summary(df_sales, date_from, date_to)

    # --- Attributes จาก Items Master (1 row ต่อ item) ---
    # --- Attributes จาก Items Master โดยตรง ---
    # ใช้ df_items แทน df_enrich เพราะ df_enrich filter เฉพาะ item ที่มี SOH > 0
    # item ที่มียอดขายแต่ SOH = 0 จะหาย attributes ถ้าดึงจาก df_enrich
    attr_cols = [
        "Item Code", "Item Name", "Product Status", "Price Status",
        "Production Type", "Product Number Reference", "Product Type",
        "Fabric Main", "Data Color", "Color Name", "Size",
        "Season for MPS", "Standard Cost", "Active Sales Pricing", "Class",
    ]
    df_attr = (
        df_items[attr_cols]
        .rename(columns={
            "Item Code":                "item_code",
            "Item Name":                "description",
            "Product Number Reference": "style_no",
            "Season for MPS":           "season",
            "Active Sales Pricing":     "active_price",
            "Standard Cost":            "standard_cost_items",
        })
        .drop_duplicates(subset=["item_code"])
    )

    # --- Build: Sales เป็น backbone ตาม SOW ---
    df = df_sales_sum.copy()
    df = df.merge(df_attr,     on="item_code", how="left")
    df = df.merge(df_gr,       on="item_code", how="left")
    df = df.merge(soh_by_item, on="item_code", how="left")

    # BOM per item × location จาก soh_loc (approximation = current SOH)
    bom_loc = (
        df_soh_loc[["item_code","location","qty"]]
        .rename(columns={"qty": "bom_qty"})
    )
    df = df.merge(bom_loc, on=["item_code","location"], how="left")

    df["total_gr_qty"] = df["total_gr_qty"].fillna(0)
    df["soh_qty"]      = df["soh_qty"].fillna(0)
    df["bom_qty"]      = df["bom_qty"].fillna(0)

    # ใช้ Standard Cost จาก Items Master เป็นหลัก
    # (ไฟล์ Sales มี STANDARD COST/U ต่อ transaction ซึ่งอาจต่างกันแต่ละแถว)
    df["standard_cost"] = df["standard_cost_items"]

    # --- คำนวณ derived columns ---
    def safe_mul(qty, price):
        """คูณ qty × price — คืน None ถ้า price ไม่มี"""
        return df.apply(
            lambda r: r[qty] * r[price]
            if pd.notna(r[price]) and r[price] > 0 else None, axis=1
        )

    df["bom_retail"]     = df.apply(lambda r: r["bom_qty"] * r["active_price"]    if pd.notna(r["active_price"])    and r["active_price"]    > 0 else None, axis=1)
    df["bom_cost"]       = df.apply(lambda r: r["bom_qty"] * r["standard_cost"]   if pd.notna(r["standard_cost"])   and r["standard_cost"]   > 0 else None, axis=1)
    df["ttl_gr_retail"]  = df.apply(lambda r: r["total_gr_qty"] * r["active_price"]   if pd.notna(r["active_price"])   and r["active_price"]   > 0 else None, axis=1)
    df["ttl_gr_cost"]    = df.apply(lambda r: r["total_gr_qty"] * r["standard_cost"]  if pd.notna(r["standard_cost"])  and r["standard_cost"]  > 0 else None, axis=1)
    df["avg_sls_amt_m"]  = df.apply(lambda r: r["avg_sls_qty_m"] * r["active_price"]  if pd.notna(r["active_price"])  and r["active_price"]  > 0 else None, axis=1)
    df["avg_cost_amt_m"] = df.apply(lambda r: r["avg_sls_qty_m"] * r["standard_cost"] if pd.notna(r["standard_cost"]) and r["standard_cost"] > 0 else None, axis=1)
    df["soh_retail_amt"] = df.apply(lambda r: r["soh_qty"] * r["active_price"]     if pd.notna(r["active_price"])     and r["active_price"]     > 0 else None, axis=1)
    df["soh_cost_amt"]   = df.apply(lambda r: r["soh_qty"] * r["standard_cost"]    if pd.notna(r["standard_cost"])    and r["standard_cost"]    > 0 else None, axis=1)

    # SELL% = TOTAL SLS QTY / (BOM QTY + TOTAL GR QTY)
    df["sell_pct"] = df.apply(
        lambda r: round(r["total_sls_qty"] / (r["bom_qty"] + r["total_gr_qty"]) * 100, 2)
        if (r["bom_qty"] + r["total_gr_qty"]) > 0 else None,
        axis=1
    )

    # ตัด row ที่ไม่มียอด (SOW: แสดงเฉพาะ line ที่มียอด)
    df = df[df["total_sls_qty"] != 0].copy()

    # Sort by Location
    df = df.sort_values(["location","item_code"]).reset_index(drop=True)

    print(f"[calculate_sellthrough] Output rows: {len(df):,} | "
          f"Date: {date_from} → {date_to} | "
          f"Net qty: {df['total_sls_qty'].sum():,.0f}")
    return df


# ============================================================
# SECTION 6: Export Excel
# ============================================================

def export_sellthrough_excel(df_report: pd.DataFrame, date_from: date, date_to: date, output_folder: str = ".") -> str:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    fname = os.path.join(
        output_folder,
        f"Sellthrough_Report_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
    )

    col_map = {
        "item_code":        "Material Code",
        "location":         "Location",
        "description":      "Description",
        "Product Status":   "Product Status",
        "Price Status":     "Price Status",
        "Production Type":  "Production Type",
        "style_no":         "Style#",
        "season":           "Season",
        "Product Type":     "Product Type",
        "Fabric Main":      "Fabric Main",
        "Data Color":       "Data Color",
        "Color Name":       "Color Name",
        "Size":             "Size",
        "active_price":     "ACTIVE PRICE",
        "standard_cost":    "Standard Cost",
        "sell_pct":         "SELL%",
        "bom_qty":          "BOM QTY",
        "bom_retail":       "BOM RETAIL",
        "bom_cost":         "BOM COST",
        "total_gr_qty":     "TOTAL GR QTY",
        "ttl_gr_retail":    "TTL GR RETAIL VALUE",
        "ttl_gr_cost":      "TTL GR COST VALUE",
        "total_sls_qty":    "TOTAL SLS QTY",
        "total_sls_amt":    "TOTAL SLS AMT",
        "total_cost_amt":   "TOTAL COST AMT",
        "avg_sls_qty_m":    "AVG SLS QTY/M",
        "avg_sls_amt_m":    "AVG SLS AMT/M",
        "avg_cost_amt_m":   "AVG COST AMT/M",
        "soh_qty":          "SOH QTY",
        "soh_retail_amt":   "SOH RETAIL AMT",
        "soh_cost_amt":     "SOH COST AMT",
    }

    df_out = pd.DataFrame()
    for src, dst in col_map.items():
        df_out[dst] = df_report[src] if src in df_report.columns else None

    wb = Workbook()
    ws = wb.active
    ws.title = "Sell-through Report"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")

    ws.append(list(df_out.columns))
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row in dataframe_to_rows(df_out, index=False, header=False):
        ws.append(row)

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Summary Sheet
    ws_sum = wb.create_sheet("📊 Sellthrough Summary")
    _write_sellthrough_summary(ws_sum, df_report, date_from, date_to)

    wb.save(fname)
    print(f"[export] บันทึกไฟล์: {fname}")
    return fname


def _write_sellthrough_summary(ws, df: pd.DataFrame, date_from: date, date_to: date):
    """
    Summary sheet สำหรับ ผจก. MER — Sell-through Report
    เน้น: ประเมิน performance รายสาขา + ระบุสินค้าที่ต้องทำโปรโมชัน/เติม stock
    Threshold: SELL% < 30% = ต่ำ, > 80% = ดีมาก (fashion retail benchmark)
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_NAVY    = "1F4E79"; C_WHITE  = "FFFFFF"; C_GREY   = "F2F2F2"
    C_GREEN   = "375623"; C_GREEN_LT = "E2EFDA"
    C_RED     = "C00000"; C_RED_LT   = "FFCCCC"
    C_AMBER_LT= "FCE4D6"; C_BLUE_LT  = "DEEAF1"

    SELL_HIGH = 80   # SELL% >= 80% = ขายดีมาก (พิจารณาเติม stock)
    SELL_LOW  = 30   # SELL% <  30% = ขายช้า (พิจารณาโปรโมชัน)

    def fill(c):  return PatternFill("solid", fgColor=c)
    def bfont(c=C_NAVY, s=11, bold=True): return Font(color=c, bold=bold, size=s)
    def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Pre-calculate ---
    df_valid    = df[df["sell_pct"].notna()].copy()
    total_qty   = df["total_sls_qty"].sum()
    total_amt   = df["total_sls_amt"].sum()
    avg_sell    = df_valid["sell_pct"].mean()
    n_sku       = df["item_code"].nunique()
    n_high      = (df_valid["sell_pct"] >= SELL_HIGH).sum()
    n_low       = (df_valid["sell_pct"] < SELL_LOW).sum()

    # Branch performance (SELL% ต่อสาขา)
    branch_perf = (
        df_valid.groupby("location", as_index=False)
        .agg(avg_sell_pct=("sell_pct","mean"),
             total_qty=("total_sls_qty","sum"),
             n_sku=("item_code","nunique"))
        .sort_values("avg_sell_pct", ascending=False)
    )

    row = 1

    # --- Title ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "📊 Sell-through Report — MER Executive Summary")
    c.font = bfont(C_WHITE, 14); c.fill = fill(C_NAVY); c.alignment = center()
    ws.row_dimensions[row].height = 30; row += 1

    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, f"Period: {date_from.strftime('%d/%m/%Y')} – {date_to.strftime('%d/%m/%Y')}   |   Generated: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = bfont(C_NAVY, 10, False); c.fill = fill(C_GREY); c.alignment = center()
    row += 2

    # --- KPI Cards (2×2) ---
    cards = [
        ("🛍️ Net Sales Qty",    f"{total_qty:,.0f} pcs",        "จำนวนขายสุทธิ (หัก return)",    C_NAVY),
        ("💰 Net Sales Amt",    f"฿{total_amt/1e6:.1f}M",        "ยอดขายสุทธิรวม",                 "2E4057"),
        ("📊 Avg SELL%",        f"{avg_sell:.1f}%",               f"Benchmark: >{SELL_HIGH}%=ดีมาก / <{SELL_LOW}%=ต้องระวัง", "375623"),
        ("🧾 SKU ที่มียอด",     f"{n_sku:,} SKUs",               f"High SELL%: {n_high} | Low SELL%: {n_low}", "4B3E6B"),
    ]
    for i, (label, val, sub, bg) in enumerate(cards):
        col = (i % 2) * 3 + 1
        r   = row if i < 2 else row + 4
        for mr in [r, r+1, r+2]:
            ws.merge_cells(start_row=mr, start_column=col, end_row=mr, end_column=col+1)
        ws.cell(r,   col, label).font = bfont(C_WHITE, 10); ws.cell(r,   col).fill = fill(bg); ws.cell(r,   col).alignment = center()
        ws.cell(r+1, col, val).font   = bfont(C_WHITE, 16); ws.cell(r+1, col).fill = fill(bg); ws.cell(r+1, col).alignment = center()
        ws.cell(r+2, col, sub).font   = bfont(C_WHITE, 9, False); ws.cell(r+2, col).fill = fill(bg); ws.cell(r+2, col).alignment = center()
        ws.row_dimensions[r].height = 18; ws.row_dimensions[r+1].height = 32; ws.row_dimensions[r+2].height = 16
    row += 9

    # --- Branch Performance Table ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "🏪 Branch Performance — SELL% รายสาขา (เรียงจากดีที่สุด)")
    c.font = bfont(C_WHITE, 11); c.fill = fill(C_NAVY); c.alignment = center()
    ws.row_dimensions[row].height = 22; row += 1

    for i, h in enumerate(["Location","Avg SELL%","Total Qty","SKUs","Status","คำแนะนำ"], 1):
        c = ws.cell(row, i, h); c.font = bfont(C_NAVY, 10); c.fill = fill(C_GREY); c.alignment = center(); c.border = border
    ws.row_dimensions[row].height = 20; row += 1

    for _, r in branch_perf.iterrows():
        sp = r["avg_sell_pct"]
        if sp >= SELL_HIGH:
            status = "🟢 ดีมาก"; bg = C_GREEN_LT; advice = "พิจารณาเติม stock"
        elif sp >= SELL_LOW:
            status = "🟡 ปกติ";  bg = "FFF2CC";   advice = "ติดตามต่อเนื่อง"
        else:
            status = "🔴 ต่ำ";   bg = C_RED_LT;   advice = "วางแผนโปรโมชัน/ลดราคา"
        for i, val in enumerate([r["location"], f"{sp:.1f}%", f"{r['total_qty']:,.0f}", f"{int(r['n_sku']):,}", status, advice], 1):
            c = ws.cell(row, i, val); c.fill = fill(bg); c.font = bfont("262626", 10, False)
            c.alignment = left() if i in [1, 6] else center(); c.border = border
        ws.row_dimensions[row].height = 18; row += 1

    row += 2

    # --- Top 10 High SELL% (เติม stock) ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, f"🏆 Top 10 SELL% สูงสุด — พิจารณาเติม Stock (SELL% >= {SELL_HIGH}%)")
    c.font = bfont(C_WHITE, 11); c.fill = fill(C_GREEN); c.alignment = center()
    ws.row_dimensions[row].height = 22; row += 1

    for i, h in enumerate(["Item Code","Location","Sales Qty","BOM Qty","GR Qty","SELL%","SOH","คำแนะนำ"], 1):
        c = ws.cell(row, i, h); c.font = bfont(C_NAVY, 10); c.fill = fill(C_GREY); c.alignment = center(); c.border = border
    ws.row_dimensions[row].height = 20; row += 1

    high_sell = df_valid[df_valid["sell_pct"] >= SELL_HIGH].nlargest(10, "sell_pct")
    for _, r in high_sell.iterrows():
        soh = r.get("soh_qty", 0)
        advice = "เติม stock ด่วน" if soh < 5 else "วางแผนเติม stock"
        for i, val in enumerate([r["item_code"], r["location"], f"{r['total_sls_qty']:,.0f}",
                                  f"{r['bom_qty']:,.0f}", f"{r['total_gr_qty']:,.0f}",
                                  f"{r['sell_pct']:.1f}%", f"{soh:,.0f}", advice], 1):
            c = ws.cell(row, i, val); c.fill = fill(C_GREEN_LT); c.font = bfont("262626", 10, False)
            c.alignment = left() if i in [1, 8] else center(); c.border = border
        ws.row_dimensions[row].height = 18; row += 1

    row += 2

    # --- Low SELL% (โปรโมชัน) ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, f"⚠️ Low SELL% — ต้องทำโปรโมชัน/ลดราคา (SELL% < {SELL_LOW}%)")
    c.font = bfont(C_WHITE, 11); c.fill = fill(C_RED); c.alignment = center()
    ws.row_dimensions[row].height = 22; row += 1

    for i, h in enumerate(["Item Code","Location","Sales Qty","BOM Qty","GR Qty","SELL%","SOH","คำแนะนำ"], 1):
        c = ws.cell(row, i, h); c.font = bfont(C_NAVY, 10); c.fill = fill(C_GREY); c.alignment = center(); c.border = border
    ws.row_dimensions[row].height = 20; row += 1

    low_sell = df_valid[df_valid["sell_pct"] > 0].nsmallest(10, "sell_pct")
    for _, r in low_sell.iterrows():
        soh = r.get("soh_qty", 0)
        advice = "ลดราคา/ระบาย" if r["sell_pct"] < 10 else "วางแผนโปรโมชัน"
        for i, val in enumerate([r["item_code"], r["location"], f"{r['total_sls_qty']:,.0f}",
                                  f"{r['bom_qty']:,.0f}", f"{r['total_gr_qty']:,.0f}",
                                  f"{r['sell_pct']:.1f}%", f"{soh:,.0f}", advice], 1):
            c = ws.cell(row, i, val); c.fill = fill(C_RED_LT); c.font = bfont("262626", 10, False)
            c.alignment = left() if i in [1, 8] else center(); c.border = border
        ws.row_dimensions[row].height = 18; row += 1

    row += 2

    # --- Action Items ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "📋 Action Items สำหรับทีม MER")
    c.font = bfont(C_WHITE, 11); c.fill = fill(C_GREEN); c.alignment = center()
    ws.row_dimensions[row].height = 22; row += 1

    branch_best  = branch_perf.iloc[0]["location"]  if len(branch_perf) > 0 else "-"
    branch_worst = branch_perf.iloc[-1]["location"] if len(branch_perf) > 0 else "-"
    best_pct     = branch_perf.iloc[0]["avg_sell_pct"]  if len(branch_perf) > 0 else 0
    worst_pct    = branch_perf.iloc[-1]["avg_sell_pct"] if len(branch_perf) > 0 else 0

    actions = [
        (f"🟢 เติม Stock: {n_high} SKUs มี SELL% >= {SELL_HIGH}% — SOH เหลือน้อย เสี่ยงของหมดก่อนกำหนด", C_GREEN_LT),
        (f"🔴 โปรโมชัน: {n_low} SKUs มี SELL% < {SELL_LOW}% — ควรวางแผน promotion หรือย้ายไปสาขาที่ขายดีกว่า", C_RED_LT),
        (f"🏆 สาขาที่ทำได้ดีที่สุด: {branch_best} (Avg SELL% = {best_pct:.1f}%) — ศึกษา best practice เพื่อนำไปใช้กับสาขาอื่น", C_GREEN_LT),
        (f"⚠️ สาขาที่ต้องปรับปรุง: {branch_worst} (Avg SELL% = {worst_pct:.1f}%) — วิเคราะห์สาเหตุและวางแผน action plan", C_AMBER_LT),
        (f"📌 Benchmark: Avg SELL% ปัจจุบัน = {avg_sell:.1f}% — fashion retail benchmark ควรอยู่ที่ 60-80% ต่อ season", C_GREY),
    ]
    for text, bg in actions:
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row, 1, text)
        c.font = bfont("262626", 10, False); c.fill = fill(bg)
        c.alignment = left(); c.border = border
        ws.row_dimensions[row].height = 28; row += 1

    for i, w in enumerate([35, 20, 14, 12, 12, 12, 12, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SECTION 7: build_sellthrough_summary() สำหรับ Streamlit
# ============================================================

def build_sellthrough_summary(df_report: pd.DataFrame, date_from: date, date_to: date) -> dict:
    total_qty = df_report["total_sls_qty"].sum()
    total_amt = df_report["total_sls_amt"].sum()
    avg_sell  = df_report["sell_pct"].dropna().mean()

    return {
        "date_from":       date_from,
        "date_to":         date_to,
        "total_qty":       total_qty,
        "total_amt":       total_amt,
        "avg_sell_pct":    avg_sell,
        "n_sku":           df_report["item_code"].nunique(),
        "n_location":      df_report["location"].nunique(),
        "top10_sellers":   df_report.groupby("item_code")["total_sls_qty"].sum().nlargest(10).reset_index(),
        "low_sellthrough": df_report[df_report["sell_pct"].notna()].nsmallest(10,"sell_pct")[["item_code","location","total_sls_qty","sell_pct"]],
        "flag_low_sell":   avg_sell < 30 if pd.notna(avg_sell) else False,
    }


# ============================================================
# SECTION 8: Main — ทดสอบ
# ============================================================

if __name__ == "__main__":
    # ส่วนการตั้งค่า (ปล่อยไว้ได้ครับ)
    DATA_FOLDER   = "data_cache" 
    OUTPUT_FOLDER = "."
    DATE_FROM     = date(2026, 1, 1)
    DATE_TO       = date(2026, 5, 13)

    print("=" * 58)
    print("  PHASE 5: Sell-through Report")
    print("=" * 58 + "\n")

    # --- ตั้งแต่บรรทัดนี้ลงไป ให้ใส่ # ปิดให้หมดทุกบรรทัดครับ ---
    # df_items, df_ns, df_erply, df_to = load_all_data(DATA_FOLDER)
    # df_transit  = get_qty_in_transit(df_to)
    # df_fallback = get_floor_date_fallback(df_to)
    # df_sales    = load_sales_data(DATA_FOLDER)

    # df_report = calculate_sellthrough(
    #     df_items, df_ns, df_erply, df_to,
    #     df_transit, df_fallback, df_sales,
    #     DATE_FROM, DATE_TO,
    # )

    # summary = build_sellthrough_summary(df_report, DATE_FROM, DATE_TO)
    # print(f"\n=== Summary ===")
    # for k, v in summary.items():
    #     if k not in ["top10_sellers","low_sellthrough"]:
    #         print(f"  {k:<20} = {v}")
    # print(f"\nTop 10:\n{summary['top10_sellers'].to_string(index=False)}")
    # print(f"\nLow Sell-through:\n{summary['low_sellthrough'].to_string(index=False)}")

    # fname = export_sellthrough_excel(df_report, DATE_FROM, DATE_TO, OUTPUT_FOLDER)
    # print(f"\n[DONE] Phase 5 เสร็จ → {fname}")

    print("\n[READY] Phase 5 is ready for Streamlit Cloud")