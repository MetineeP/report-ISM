"""
Phase 6: Accumulated Sales Report
===================================
แสดงยอดขายสะสมแยกตามเดือน Jan-Dec
เพื่อเห็น "การเติบโตและแนวโน้มของยอดขาย" ในภาพรวม

Columns ตาม SOW:
  Meta: Material Code, Location, Description, Product Status, Price Status,
        Production Type, Style#, Product Type, Fabric Main, Collection,
        Color Group, Data Color, Color Name, Size, Season for MPS
  Monthly: Jan, Feb, Mar, ... Dec (Quantity)
  Summary: TTL, AVG.M, SOH, MH

Logic:
  - TTL    = SUM(Jan-Dec)
  - AVG.M  = เฉลี่ยเฉพาะเดือนที่มียอด ≠ 0
  - SOH    = SOH QTY (NS + Erply ตาม Phase 2-3)
  - MH     = SOH / AVG.M  (Month of Hand — กี่เดือนจึงจะขายหมด)

Base: YMFSALESDATAWITHCOSTResults (ตาม SOW)
ตัดออก: Product Status = PREMIUM, SEMI
"""

import os
import sys
from datetime import date
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from phase2_soh import (
    load_all_data, get_qty_in_transit, get_floor_date_fallback,
    calculate_soh, EXCLUDE_PRODUCT_STATUS,
)
from phase3_enrich import build_enriched_dataset, build_soh_by_location
from phase5_sellthrough import load_sales_data, ERROR_SENTINEL


# ============================================================
# SECTION 1: Calculate Accumulated Sales
# ============================================================

def calculate_accumulated(
    df_items:    pd.DataFrame,
    df_ns:       pd.DataFrame,
    df_erply:    pd.DataFrame,
    df_transit:  pd.DataFrame,
    df_fallback: pd.DataFrame,
    df_sales:    pd.DataFrame,
    date_from:   date,
    date_to:     date,
) -> pd.DataFrame:
    """
    Build Accumulated Sales Report — 1 row ต่อ item × location
    Monthly columns: Jan-Dec (หรือเฉพาะเดือนที่อยู่ใน date range)

    กรณีเรียกข้ามปี เช่น Jan24-May25:
      - Jan-May แสดงยอดรวม ปี 24+25
      - Jun-Dec แสดงยอดเฉพาะปี 24 (ตาม SOW)
    """

    # --- Filter date range ---
    mask = (
        (df_sales["Selling Date"] >= pd.Timestamp(date_from)) &
        (df_sales["Selling Date"] <= pd.Timestamp(date_to))
    )
    df = df_sales[mask].copy()

    # ตัด Product Status: PREMIUM, SEMI (ตาม SOW condition 4)
    df = df[~df["Product Status"].isin(EXCLUDE_PRODUCT_STATUS)].copy()

    # --- Pivot ยอดขาย qty ต่อเดือน ---
    # Month label: "Jan", "Feb", ... "Dec"
    MONTH_LABELS = {
        1:"Jan", 2:"Feb", 3:"Mar", 4:"Apr", 5:"May", 6:"Jun",
        7:"Jul", 8:"Aug", 9:"Sep", 10:"Oct", 11:"Nov", 12:"Dec"
    }
    df["month_label"] = df["Month Sold"].astype(int).map(MONTH_LABELS)

    # กรณีข้ามปี: รวม qty ของเดือนเดียวกันทุกปีเข้าด้วยกัน
    # (Jan24 + Jan25 รวมกันใน column "Jan")
    pivot = (
        df.groupby(["Material Code", "Location Name", "month_label"], as_index=False)
        ["Quantity"].sum()
    )

    # Pivot เป็น wide format
    df_wide = pivot.pivot_table(
        index=["Material Code", "Location Name"],
        columns="month_label",
        values="Quantity",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    df_wide.columns.name = None
    df_wide = df_wide.rename(columns={
        "Material Code": "item_code",
        "Location Name": "location",
    })

    # เพิ่ม month columns ที่ไม่มีในข้อมูล → 0
    for lbl in MONTH_LABELS.values():
        if lbl not in df_wide.columns:
            df_wide[lbl] = 0

    # เรียงลำดับ month columns ตามปฏิทิน
    month_cols = list(MONTH_LABELS.values())

    # --- คำนวณ TTL และ AVG.M ---
    df_wide["TTL"] = df_wide[month_cols].sum(axis=1)

    # AVG.M = เฉลี่ยเฉพาะเดือนที่มียอด ≠ 0 (ตาม SOW)
    def avg_nonzero(row):
        vals = [row[m] for m in month_cols if row[m] != 0]
        return sum(vals) / len(vals) if vals else 0

    df_wide["AVG.M"] = df_wide.apply(avg_nonzero, axis=1)

    # --- Merge SOH ---
    df_soh     = calculate_soh(df_items, df_ns, df_erply, df_transit)
    df_enrich  = build_enriched_dataset(df_items, df_soh, df_fallback)
    df_soh_loc = build_soh_by_location(df_items, df_ns, df_erply)

    # SOH per item (รวมทุก location)
    soh_by_item = (
        df_soh_loc.groupby("item_code", as_index=False)["qty"]
        .sum().rename(columns={"qty": "SOH"})
    )
    df_wide = df_wide.merge(soh_by_item, on="item_code", how="left")
    df_wide["SOH"] = df_wide["SOH"].fillna(0)

    # MH = SOH / AVG.M (กี่เดือนจึงจะขายหมด)
    df_wide["MH"] = df_wide.apply(
        lambda r: round(r["SOH"] / r["AVG.M"], 2) if r["AVG.M"] > 0 else None,
        axis=1
    )

    # --- Merge attributes จาก Items Master ---
    # --- Attributes จาก Items Master โดยตรง ---
    # ใช้ df_items แทน df_enrich เพราะ df_enrich filter เฉพาะ item ที่มี SOH > 0
    # item ที่มียอดขายแต่ SOH = 0 จะหาย attributes ถ้าดึงจาก df_enrich
    attr_cols = [
        "Item Code", "Item Name", "Product Status", "Price Status",
        "Production Type", "Product Number Reference", "Product Type",
        "Fabric Main", "Data Color", "Color Name", "Size",
        "Season for MPS", "Class",
    ]
    df_attr = (
        df_items[attr_cols]
        .rename(columns={
            "Item Code":                "item_code",
            "Item Name":                "description",
            "Product Number Reference": "style_no",
            "Season for MPS":           "season",
        })
        .drop_duplicates(subset=["item_code"])
    )
    df_wide = df_wide.merge(df_attr, on="item_code", how="left")

    # --- ตัด row ที่ TTL = 0 (SOW: แสดงเฉพาะ line ที่มียอด) ---
    df_wide = df_wide[df_wide["TTL"] != 0].copy()

    # --- Sort by Location ---
    df_wide = df_wide.sort_values(["location", "item_code"]).reset_index(drop=True)

    print(
        f"[calculate_accumulated] Output rows: {len(df_wide):,} | "
        f"Date: {date_from} → {date_to} | "
        f"TTL qty: {df_wide['TTL'].sum():,.0f}"
    )
    return df_wide


# ============================================================
# SECTION 2: Export Excel
# ============================================================

def export_accumulated_excel(
    df_report:     pd.DataFrame,
    date_from:     date,
    date_to:       date,
    output_folder: str = ".",
) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    fname = os.path.join(
        output_folder,
        f"Accumulated_Sales_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
    )

    MONTH_LABELS = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]

    # เรียงลำดับ columns ตาม SOW
    meta_cols = [
        ("item_code",       "Material Code"),
        ("location",        "Location"),
        ("description",     "Description"),
        ("Product Status",  "Product Status"),
        ("Price Status",    "Price Status"),
        ("Production Type", "Production Type"),
        ("style_no",        "Style#"),
        ("Product Type",    "Product Type"),
        ("Fabric Main",     "Fabric Main"),
        ("Data Color",      "Data Color"),
        ("Color Name",      "Color Name"),
        ("Size",            "Size"),
        ("season",          "Season for MPS"),
    ]

    df_out = pd.DataFrame()
    for src, dst in meta_cols:
        df_out[dst] = df_report[src] if src in df_report.columns else None

    # Monthly columns
    for m in MONTH_LABELS:
        df_out[m] = df_report[m] if m in df_report.columns else 0

    # Summary columns
    df_out["TTL"]   = df_report["TTL"]
    df_out["AVG.M"] = df_report["AVG.M"].round(2)
    df_out["SOH"]   = df_report["SOH"]
    df_out["MH"]    = df_report["MH"]

    # เขียน Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Accumulated Sales"

    C_NAVY   = "1F4E79"
    C_TEAL   = "1F6B75"   # สีต่างสำหรับ monthly columns
    C_PURPLE = "4B3E6B"   # สีสำหรับ summary columns

    hdr_fill    = PatternFill("solid", fgColor=C_NAVY)
    hdr_fill_m  = PatternFill("solid", fgColor=C_TEAL)
    hdr_fill_s  = PatternFill("solid", fgColor=C_PURPLE)
    hdr_font    = Font(bold=True, color="FFFFFF")

    ws.append(list(df_out.columns))
    for i, cell in enumerate(ws[1]):
        col_name = cell.value
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_name in MONTH_LABELS:
            cell.fill = hdr_fill_m   # Teal สำหรับ monthly
        elif col_name in ["TTL", "AVG.M", "SOH", "MH"]:
            cell.fill = hdr_fill_s   # Purple สำหรับ summary
        else:
            cell.fill = hdr_fill     # Navy สำหรับ meta

    for row_data in dataframe_to_rows(df_out, index=False, header=False):
        ws.append(row_data)

    ws.freeze_panes = "A2"

    # Highlight negative cells (return มากกว่าขาย)
    red_font = Font(color="C00000", bold=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = red_font

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 20)

    # Summary Sheet
    ws_sum = wb.create_sheet("📊 Accumulated Summary")
    _write_accumulated_summary(ws_sum, df_report, date_from, date_to)

    wb.save(fname)
    print(f"[export] บันทึกไฟล์: {fname}")
    return fname


def _write_accumulated_summary(ws, df: pd.DataFrame, date_from: date, date_to: date):
    """
    Summary sheet สำหรับ ผจก. MER — Accumulated Sales Report
    เน้น: แนวโน้มยอดขายรายเดือน + ระบุสินค้า slow moving (MH สูง)
    Threshold MH: > 3 เดือน = เริ่มระวัง, > 6 เดือน = slow moving ชัดเจน
    (fashion retail best practice — ทีม MER ปรับได้ใน config)
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_NAVY    = "1F4E79"; C_WHITE = "FFFFFF"; C_GREY   = "F2F2F2"
    C_PURPLE_LT = "EAE4F7"; C_GREEN_LT = "E2EFDA"
    C_AMBER_LT  = "FCE4D6"; C_RED_LT   = "FFCCCC"
    C_GREEN   = "375623"; C_RED = "C00000"

    MH_WARN     = 3   # MH > 3 เดือน = เริ่มระวัง
    MH_CRITICAL = 6   # MH > 6 เดือน = slow moving

    def fill(c):  return PatternFill("solid", fgColor=c)
    def bfont(c=C_NAVY, s=11, bold=True): return Font(color=c, bold=bold, size=s)
    def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    MONTH_LABELS = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]

    # --- Pre-calculate ---
    total_qty      = df["TTL"].sum()
    avg_mh         = df["MH"].dropna().mean()
    n_sku          = df["item_code"].nunique()
    n_location     = df["location"].nunique()
    monthly_totals = {m: df[m].sum() for m in MONTH_LABELS if m in df.columns}
    peak_month     = max(monthly_totals, key=monthly_totals.get) if any(v > 0 for v in monthly_totals.values()) else "-"
    n_warn         = (df["MH"] > MH_WARN).sum()
    n_critical     = (df["MH"] > MH_CRITICAL).sum()

    row = 1

    # --- Title ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "📊 Accumulated Sales Report — MER Executive Summary")
    c.font = bfont(C_WHITE, 14); c.fill = fill(C_NAVY); c.alignment = center()
    ws.row_dimensions[row].height = 30; row += 1

    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, f"Period: {date_from.strftime('%d/%m/%Y')} – {date_to.strftime('%d/%m/%Y')}   |   Generated: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = bfont(C_NAVY, 10, False); c.fill = fill(C_GREY); c.alignment = center()
    row += 2

    # --- KPI Cards ---
    cards = [
        ("📦 Total Qty Sold",   f"{total_qty:,.0f} pcs",       "ยอดขายรวมทุกเดือน",                C_NAVY),
        ("📊 Avg MH",           f"{avg_mh:.1f} เดือน",         f">{MH_WARN}M=ระวัง / >{MH_CRITICAL}M=slow moving", "2E4057"),
        ("🏔️ Peak Month",       peak_month,                     f"เดือนยอดขายสูงสุด = {monthly_totals.get(peak_month,0):,.0f} pcs", "375623"),
        ("⚠️ Slow Moving SKUs", f"{n_critical:,} SKUs",         f"MH > {MH_CRITICAL} เดือน (ระวัง {n_warn:,} SKUs ที่ MH > {MH_WARN})", "C00000"),
    ]
    for i, (label, val, sub, bg) in enumerate(cards):
        col = (i % 2) * 3 + 1
        r   = row if i < 2 else row + 4
        for mr in [r, r+1, r+2]:
            ws.merge_cells(start_row=mr, start_column=col, end_row=mr, end_column=col+1)
        ws.cell(r,   col, label).font = bfont(C_WHITE, 10); ws.cell(r,   col).fill = fill(bg); ws.cell(r,   col).alignment = center()
        ws.cell(r+1, col, val).font   = bfont(C_WHITE, 16); ws.cell(r+1, col).fill = fill(bg); ws.cell(r+1, col).alignment = center()
        ws.cell(r+2, col, sub).font   = bfont(C_WHITE, 9, False); ws.cell(r+2, col).fill = fill(bg); ws.cell(r+2, col).alignment = center()
        ws.row_dimensions[r].height = 18; ws.row_dimensions[r+1].height = 32; ws.row_dimensions[r+2].height = 16
    row += 9

    # --- Monthly Trend Table ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "📅 Monthly Sales Trend — แนวโน้มยอดขายรายเดือน")
    c.font = bfont(C_WHITE, 11); c.fill = fill(C_NAVY); c.alignment = center()
    ws.row_dimensions[row].height = 22; row += 1

    for i, h in enumerate(["Month","Qty","% of Total","Growth MoM","สถานะ"], 1):
        c = ws.cell(row, i, h); c.font = bfont(C_NAVY, 10); c.fill = fill(C_GREY); c.alignment = center(); c.border = border
    ws.row_dimensions[row].height = 20; row += 1

    max_qty  = max(monthly_totals.values()) if monthly_totals else 1
    prev_qty = None
    for m in MONTH_LABELS:
        qty  = monthly_totals.get(m, 0)
        pct  = qty / total_qty * 100 if total_qty != 0 else 0
        mom  = f"{(qty-prev_qty)/prev_qty*100:+.1f}%" if prev_qty and prev_qty > 0 and qty > 0 else "-"
        if qty == 0:
            status = "-"; bg = C_GREY
        elif qty == max_qty:
            status = "🏔️ Peak"; bg = C_PURPLE_LT
        elif prev_qty and qty > prev_qty:
            status = "📈 ขึ้น"; bg = C_GREEN_LT
        elif prev_qty and qty < prev_qty:
            status = "📉 ลง";  bg = C_RED_LT
        else:
            status = "➡️ คงที่"; bg = C_GREY

        for i, val in enumerate([m, f"{qty:,.0f}" if qty > 0 else "-", f"{pct:.1f}%" if qty > 0 else "-", mom, status], 1):
            c = ws.cell(row, i, val); c.fill = fill(bg); c.font = bfont("262626", 10, False)
            c.alignment = center(); c.border = border
        ws.row_dimensions[row].height = 18; row += 1
        if qty > 0: prev_qty = qty

    # Total row
    for i, val in enumerate(["TOTAL", f"{total_qty:,.0f}", "100.0%", "", ""], 1):
        c = ws.cell(row, i, val); c.fill = fill(C_NAVY); c.font = bfont(C_WHITE, 10)
        c.alignment = center(); c.border = border
    ws.row_dimensions[row].height = 20; row += 2

    # --- Top 10 Best Sellers ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "🏆 Top 10 Best Sellers — ยอดสะสมสูงสุด")
    c.font = bfont(C_WHITE, 11); c.fill = fill(C_GREEN); c.alignment = center()
    ws.row_dimensions[row].height = 22; row += 1

    for i, h in enumerate(["Item Code","TTL Qty","AVG.M/เดือน","SOH","MH","% of Total","สถานะ MH","คำแนะนำ"], 1):
        c = ws.cell(row, i, h); c.font = bfont(C_NAVY, 10); c.fill = fill(C_GREY); c.alignment = center(); c.border = border
    ws.row_dimensions[row].height = 20; row += 1

    top10 = df.groupby("item_code").agg(
        ttl=("TTL","sum"), avg_m=("AVG.M","mean"), soh=("SOH","sum"), mh=("MH","mean")
    ).nlargest(10,"ttl").reset_index()

    for _, r in top10.iterrows():
        pct  = r["ttl"] / total_qty * 100 if total_qty > 0 else 0
        mh_v = r["mh"]
        if pd.isna(mh_v): mh_status, advice, bg = "N/A", "-", C_GREY
        elif mh_v > MH_CRITICAL: mh_status, advice, bg = f"🔴 {mh_v:.1f}M", "วางแผนระบาย", C_RED_LT
        elif mh_v > MH_WARN:     mh_status, advice, bg = f"🟡 {mh_v:.1f}M", "เฝ้าระวัง", C_AMBER_LT
        else:                     mh_status, advice, bg = f"🟢 {mh_v:.1f}M", "ปกติ", C_GREEN_LT

        for i, val in enumerate([r["item_code"], f"{r['ttl']:,.0f}", f"{r['avg_m']:.1f}", f"{r['soh']:,.0f}", mh_status, f"{pct:.1f}%", mh_status, advice], 1):
            c = ws.cell(row, i, val); c.fill = fill(bg); c.font = bfont("262626", 10, False)
            c.alignment = left() if i in [1, 8] else center(); c.border = border
        ws.row_dimensions[row].height = 18; row += 1

    row += 2

    # --- Slow Moving (MH สูง) ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, f"⚠️ Slow Moving — MH > {MH_CRITICAL} เดือน (ขายช้า ของจะค้าง)")
    c.font = bfont(C_WHITE, 11); c.fill = fill(C_RED); c.alignment = center()
    ws.row_dimensions[row].height = 22; row += 1

    for i, h in enumerate(["Item Code","Location","TTL Qty","AVG.M","SOH","MH","สถานะ","คำแนะนำ"], 1):
        c = ws.cell(row, i, h); c.font = bfont(C_NAVY, 10); c.fill = fill(C_GREY); c.alignment = center(); c.border = border
    ws.row_dimensions[row].height = 20; row += 1

    high_mh = df[df["MH"] > MH_CRITICAL].nlargest(10,"MH")[["item_code","location","TTL","AVG.M","SOH","MH"]]
    for _, r in high_mh.iterrows():
        mh_v    = r["MH"]
        status  = f"🔴 {mh_v:.1f}M"
        advice  = "ลดราคา/ย้ายสาขา" if mh_v > 12 else "วางแผนโปรโมชัน"
        for i, val in enumerate([r["item_code"], r["location"], f"{r['TTL']:,.0f}",
                                  f"{r['AVG.M']:.1f}", f"{r['SOH']:,.0f}", f"{mh_v:.1f}", status, advice], 1):
            c = ws.cell(row, i, val); c.fill = fill(C_RED_LT); c.font = bfont("262626", 10, False)
            c.alignment = left() if i in [1, 2, 8] else center(); c.border = border
        ws.row_dimensions[row].height = 18; row += 1

    row += 2

    # --- Action Items ---
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "📋 Action Items สำหรับทีม MER")
    c.font = bfont(C_WHITE, 11); c.fill = fill(C_GREEN); c.alignment = center()
    ws.row_dimensions[row].height = 22; row += 1

    # หาเดือนที่ลดลงมากที่สุด
    month_list  = [(m, monthly_totals.get(m, 0)) for m in MONTH_LABELS if monthly_totals.get(m, 0) > 0]
    drop_months = []
    for i in range(1, len(month_list)):
        prev, curr = month_list[i-1][1], month_list[i][1]
        if prev > 0 and curr < prev:
            drop_pct = (prev - curr) / prev * 100
            drop_months.append((month_list[i][0], drop_pct))
    worst_drop = max(drop_months, key=lambda x: x[1]) if drop_months else None

    actions = [
        (f"🔴 Slow Moving: {n_critical:,} SKUs มี MH > {MH_CRITICAL} เดือน — ควรวางแผนโปรโมชันหรือย้ายสาขาเพื่อระบาย", C_RED_LT),
        (f"🟡 เฝ้าระวัง: {n_warn:,} SKUs มี MH > {MH_WARN} เดือน — ติดตามอย่างใกล้ชิด ก่อนกลายเป็น slow moving", C_AMBER_LT),
        (f"🏔️ Peak Month: {peak_month} มียอดขายสูงสุด {monthly_totals.get(peak_month,0):,.0f} pcs — วางแผนเตรียม stock ล่วงหน้าสำหรับปีถัดไป", C_GREEN_LT),
        (f"📉 เดือนที่ยอดลดมากที่สุด: {worst_drop[0]} (-{worst_drop[1]:.1f}% MoM) — วิเคราะห์สาเหตุและวางแผน promotion" if worst_drop else "📉 ยังไม่มีข้อมูลเพียงพอสำหรับวิเคราะห์แนวโน้ม MoM", C_AMBER_LT),
        (f"📌 Benchmark: Avg MH = {avg_mh:.1f} เดือน — fashion retail ควรอยู่ที่ < 3 เดือน ถ้าสูงกว่านี้ควรทบทวนแผนการผลิต", C_GREY),
    ]
    for text, bg in actions:
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row, 1, text)
        c.font = bfont("262626", 10, False); c.fill = fill(bg)
        c.alignment = left(); c.border = border
        ws.row_dimensions[row].height = 28; row += 1

    for i, w in enumerate([35, 25, 14, 12, 12, 12, 14, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SECTION 3: build_accumulated_summary() สำหรับ Streamlit
# ============================================================

def build_accumulated_summary(df_report: pd.DataFrame, date_from: date, date_to: date) -> dict:
    """คืน dict สำหรับ Streamlit dashboard cards"""
    MONTH_LABELS = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]

    total_qty      = df_report["TTL"].sum()
    avg_mh         = df_report["MH"].dropna().mean()
    monthly_totals = {m: df_report[m].sum() for m in MONTH_LABELS if m in df_report.columns}
    peak_month     = max(monthly_totals, key=monthly_totals.get) if monthly_totals else None

    top10 = (
        df_report.groupby("item_code")
        .agg(ttl=("TTL","sum"), avg_m=("AVG.M","mean"), mh=("MH","mean"))
        .nlargest(10,"ttl").reset_index()
    )
    slow_moving = (
        df_report[df_report["MH"].notna()]
        .nlargest(10,"MH")[["item_code","location","TTL","AVG.M","SOH","MH"]]
    )

    return {
        "date_from":       date_from,
        "date_to":         date_to,
        "total_qty":       total_qty,
        "avg_mh":          avg_mh,
        "n_sku":           df_report["item_code"].nunique(),
        "n_location":      df_report["location"].nunique(),
        "monthly_totals":  monthly_totals,
        "peak_month":      peak_month,
        "top10_sellers":   top10,
        "slow_moving":     slow_moving,
        "flag_high_mh":    avg_mh > 6 if pd.notna(avg_mh) else False,  # > 6 เดือน = ควรระวัง
    }


# ============================================================
# SECTION 4: Main — ทดสอบ
# ============================================================

if __name__ == "__main__":
    # เปลี่ยนให้มาชี้ที่โฟลเดอร์ชั่วคราวที่ app.py เตรียมไว้ให้
    DATA_FOLDER   = "data_cache" 
    OUTPUT_FOLDER = "."
    DATE_FROM     = date(2026, 1, 1)
    DATE_TO       = date(2026, 5, 13)

    print("=" * 58)
    print("  PHASE 6: Accumulated Sales Report")
    print("=" * 58 + "\n")

    df_items, df_ns, df_erply, df_to = load_all_data(DATA_FOLDER)
    df_transit  = get_qty_in_transit(df_to)
    df_fallback = get_floor_date_fallback(df_to)
    df_sales    = load_sales_data(DATA_FOLDER)

    df_report = calculate_accumulated(
        df_items, df_ns, df_erply,
        df_transit, df_fallback, df_sales,
        DATE_FROM, DATE_TO,
    )

    summary = build_accumulated_summary(df_report, DATE_FROM, DATE_TO)
    print(f"\n=== Summary ===")
    for k, v in summary.items():
        if k not in ["top10_sellers","slow_moving","monthly_totals"]:
            print(f"  {k:<20} = {v}")
    print(f"\nMonthly Totals:")
    for m, qty in summary["monthly_totals"].items():
        bar = "█" * int(qty / max(summary["monthly_totals"].values()) * 20) if max(summary["monthly_totals"].values()) > 0 else ""
        print(f"  {m}: {qty:>8,.0f}  {bar}")
    print(f"\nTop 10:\n{summary['top10_sellers'].to_string(index=False)}")
    print(f"\nSlow Moving (High MH):\n{summary['slow_moving'].to_string(index=False)}")

    fname = export_accumulated_excel(df_report, DATE_FROM, DATE_TO, OUTPUT_FOLDER)
    print(f"\n[DONE] Phase 6 เสร็จ → {fname}")
